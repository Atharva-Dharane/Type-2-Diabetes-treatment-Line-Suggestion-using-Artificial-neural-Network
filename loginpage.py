# -*- coding: utf-8 -*-
"""
Created on Tue Sep  1 16:50:06 2020

@author: Shirish
"""

import PySimpleGUI as sg
import pandas as pd
from keras.models import load_model
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

def sigmoidfunc(X):
    return 1/(1+np.exp(-X))


data = pd.read_excel("UserLogins.xlsx")

model = load_model("final_ann_model.h5")

def get_patient_details(a):#dataset = pd.read_csv("DiabetesDataset.csv")
    patients = pd.read_excel("AllPatients.xlsx")
    medicines = ["Metaformin", "SGLT-2", "GLP-1", "DPP-4", "Thiazolidinediones", "Sulfonylureas", "A-Glucosesidase"]
    medicines_plt = ["Met", "SGLT-2", "GLP-1", "DPP-4", "Thiazo", "Sulfo", "A-Gluco"]
    predict1 = patients.iloc[int(a), 2:27].values
    predict = np.array([predict1])
    res1 = model.predict(predict)

    finalres1 = sigmoidfunc(res1)
    fresult = finalres1.ravel().tolist()
    #finalres2 = softmaxfunc(res1)
    d1 = dict(enumerate(finalres1.flatten(),1))

    d2 = dict(enumerate(medicines, 1))

    maxprob = np.amax(finalres1)
    for key, value in d1.items():
        if(value == maxprob):
            finalmed = key
            print(finalmed)
    pname = patients.iloc[int(a), 0]
    pID = patients.iloc[int(a), 1]
    OP_TEXT1 = "Patient ID : "+str(pID)+"              Patient Name : "+pname
    for key,value in d2.items():
        if(key == finalmed):
            medsuggestion = value
    OP_TEXT = "Suggested Medicine is : " + medsuggestion
    fig = "Figure"+str(pID)+".png"
    plt.title("Medicine-Probability Graph")
    plt.plot(medicines_plt, fresult, 'o', linestyle='solid', markerfacecolor='blue', color='orange')
    plt.annotate('Medicine Suggested', (medsuggestion,maxprob))
    plt.xlabel('Medicines')
    plt.ylabel('Probabilities')
    plt.savefig(fig)
    plt.clf()
    wb = openpyxl.load_workbook("AllPatients.xlsx")
    r = wb.active
    cellup = r.cell(row = int(a)+2,column=28)
    cellup.value = medsuggestion
    cellgraph = r.cell(row = int(a)+2,column=29)
    cellgraph.value = fig
    print("BP14 in register",a)
    wb.save("AllPatients.xlsx") 
    #print("After open image")
    patient_detail_layout=[[sg.T(OP_TEXT1,key = "name")],[sg.T(OP_TEXT,key = "OP")],
                            [sg.Image(fig)],
                            [sg.B("OK",key = "okclose"),sg.B("Delete",key = "delete"),sg.B("Edit",key = "edit")]]
    patient_window = sg.Window("Details",patient_detail_layout)
    
    event, values = patient_window.Read()
    print("BP22 bfr okclose",a)
    while event!="okclose":   
        print("BP21 in not okclose event",event)        # Event Loop
        print("BP23 in not okclose event",event)
        if event=="delete":
            print("BP20 in delete",a)
            wb = openpyxl.load_workbook("AllPatients.xlsx")
            r = wb.active
            print("BP23 in not okclose event",event) 
            r.delete_rows(int(a)+2)
            wb.save("AllPatients.xlsx") 
            break
        elif event=="edit":
            edit(int(a)+2,pname,pID,predict)
            break
        event, values = patient_window.Read()
    patient_window.close()
    go_to_home()

def edit(n,name,pID,predict):
    print("BP1 EDIT patient number ",n)
    print("BP2 EDIT patient predict ",predict)
    old_val =[]
    for i in range(25):
        k = predict[0][i]
        print(" i ",i,"atr,")
        if k == 1:
            k = True
        elif k == 0:
            k = False
        old_val.insert(i,k)
        
    edit_patient_layout=[[sg.T("Patient's name : "),sg.T(name),sg.T("Patient's ID : "),sg.T(pID)],
                         [sg.T("Enter Patient's HBA1C Level : "),sg.I(old_val[0],key="Atr1")],
                         [sg.T("Select weight : "),sg.Radio("Underweight",default=old_val[1],key="Atr2",group_id = 1),sg.Radio("Healthy",key="Atr3",default=old_val[2],group_id = 1),sg.Radio("Overweight",key="Atr4",group_id = 1,default=old_val[3]),sg.Radio("Obese",key="Atr5",group_id = 1,default=old_val[4])],
                         [sg.Checkbox("Pregnant?   ",key="Atr6",default=old_val[6])],[sg.Checkbox("Hypothyroid",key="Atr7",default=old_val[7])],
                         [sg.CBox("Genital Itching ",key="Atr8",default=old_val[8])],[sg.CBox("Gastrointestinal Problems",key="Atr9",default=old_val[9])],
                         [sg.CBox("Lactic Acidosis ",key="Atr10",default=old_val[10])],[sg.CBox("Liver Disease",key="Atr11",default=old_val[11])],
                         [sg.CBox("Kidney Disease  ",key="Atr12",default=old_val[12])],[sg.CBox("Cardiovascular Problems",key="Atr13",default=old_val[13])],
                         [sg.CBox("Pancreatis      ",key="Atr14",default=old_val[14])],[sg.CBox("Nasopharyngitis",key="Atr15",default=old_val[15])],
                         [sg.CBox("Oedema          ",key="Atr16",default=old_val[16])],[sg.CBox("Hypotension",key="Atr17",default=old_val[17])],
                         [sg.CBox("LDL             ",key="Atr18",default=old_val[18])],[sg.CBox("Cell Tumour",key="Atr19",default=old_val[19])],
                         [sg.CBox("Bladder Cancer  ",key="Atr20",default=old_val[20])],[sg.CBox("B12 Deficiency",key="Atr21",default=old_val[21])],
                         [sg.CBox("Gangrene        ",key="Atr22",default=old_val[22])],[sg.CBox("Fracture",key="Atr23",default=old_val[23])],
                         [sg.CBox("Financial Issues",key="Atr24",default=old_val[24])],[sg.CBox("Hypoglycemia",key="Atr25",default=old_val[5])],
                         [sg.B("Save",key="Save")]]
    edit_patient_window = sg.Window("Add Patient",edit_patient_layout)
    event, values = edit_patient_window.Read()
    atr_list = ["add_name","add_ID","Atr1","Atr2","Atr3","Atr4","Atr5","Atr25","Atr6","Atr7","Atr8","Atr9","Atr10","Atr11","Atr12","Atr13","Atr14","Atr15","Atr16","Atr17","Atr18","Atr19","Atr20","Atr21","Atr22","Atr23","Atr24",]
    flag = 0
    while event!="Save":
        event, values = edit_patient_window.Read()    
        if event in (None, 'Exit'):
            flag = 1
            break
    if flag == 0:
        i=2
        ip = []
        ip.insert(0,name)
        ip.insert(1,pID)
        while i<len(atr_list):
            k = values[atr_list[i]]
            print(" i ",i,"atr,",atr_list[i])
            if k == True:
                k = 1
            elif k == False:
                k = 0
            ip.insert(i,k)
            i = i+1
        ip.insert(len(atr_list),"NA")
        ip.insert(len(atr_list)+1,"NA")
        wb = openpyxl.load_workbook("AllPatients.xlsx")
        print("BP 20")
        r = wb.active
        print("BP 21")
        for i in range(3,28):
            temp = r.cell(row = n,column=i)
            temp.value = ip[i-1]
        print("BP 23 after delete row")
        wb.save("AllPatients.xlsx")
    edit_patient_window.close()
    """wb = openpyxl.load_workbook("AllPatients.xlsx")
    r = wb.active
    print("BP23 in not okclose") 
    ip = r.rows(n)
    print("inside edit bp2",ip)
    wb.save("AllPatients.xlsx")""" 
    
def new_patient():
    add_patient_layout=[[sg.T("Enter Patient's name : "),sg.I(key="add_name"),sg.T("Enter Patient's ID : "),sg.I(key="add_ID"),sg.T("ID is not unique",key="un_ID",visible=False)],
                         [sg.T("Enter Patient's HBA1C Level : "),sg.I(key="Atr1")],
                         [sg.T("Select weight : "),sg.Radio("Underweight",key="Atr2",group_id = 1),sg.Radio("Healthy",key="Atr3",group_id = 1),sg.Radio("Overweight",key="Atr4",group_id = 1),sg.Radio("Obese",key="Atr5",group_id = 1)],
                         [sg.Checkbox("Pregnant?   ",key="Atr6")],[sg.Checkbox("Hypothyroid",key="Atr7")],
                         [sg.CBox("Genital Itching ",key="Atr8")],[sg.CBox("Gastrointestinal Problems",key="Atr9")],
                         [sg.CBox("Lactic Acidosis ",key="Atr10")],[sg.CBox("Liver Disease",key="Atr11")],
                         [sg.CBox("Kidney Disease  ",key="Atr12")],[sg.CBox("Cardiovascular Problems",key="Atr13")],
                         [sg.CBox("Pancreatis      ",key="Atr14")],[sg.CBox("Nasopharyngitis",key="Atr15")],
                         [sg.CBox("Oedema          ",key="Atr16")],[sg.CBox("Hypotension",key="Atr17")],
                         [sg.CBox("LDL             ",key="Atr18")],[sg.CBox("Cell Tumour",key="Atr19")],
                         [sg.CBox("Bladder Cancer  ",key="Atr20")],[sg.CBox("B12 Deficiency",key="Atr21")],
                         [sg.CBox("Gangrene        ",key="Atr22")],[sg.CBox("Fracture",key="Atr23")],
                         [sg.CBox("Financial Issues",key="Atr24")],[sg.CBox("Hypoglycemia",key="Atr25")],
                         [sg.B("Submit",key="Submit")]]
    new_patient_window = sg.Window("Add Patient",add_patient_layout)
    event, values = new_patient_window.Read()
    atr_list = ["add_name","add_ID","Atr1","Atr2","Atr3","Atr4","Atr5","Atr25","Atr6","Atr7","Atr8","Atr9","Atr10","Atr11","Atr12","Atr13","Atr14","Atr15","Atr16","Atr17","Atr18","Atr19","Atr20","Atr21","Atr22","Atr23","Atr24",]
    while True:
        event, values = new_patient_window.Read()
        if event in (None, 'Exit'):
            break
        if event =="Submit":
            valid = verify_id(values["add_ID"])
            if valid == 1:
                new_patient_window["un_ID"].update(visible = True)
            else:
                i=0
                ip = []
                while i<len(atr_list):
                    k = values[atr_list[i]]
                    print(" i ",i,"atr,",atr_list[i])
                    if k == True:
                        k = 1
                    elif k == False:
                        k = 0
                    ip.insert(i,k)
                    i = i+1
                ip.insert(len(atr_list),"NA")
                ip.insert(len(atr_list)+1,"NA")

                print("IP",ip,len(ip))
                wb = openpyxl.load_workbook("AllPatients.xlsx")
                r = wb.active
                r.append(ip)
                print("BP 18")
                wb.save("AllPatients.xlsx")
                print("BP 19")
                print("BP 15 IP",ip)
                patients = pd.read_excel("AllPatients.xlsx")
                print("DATAFRAME PATIENTS",patients)
                new_patient_window.close()
    go_to_home()
    print("Here BP13")
def go_to_home():
    patients = pd.read_excel("AllPatients.xlsx")
    all_patients = patients[:].values.tolist()
    homepage_layout = [[sg.Table(headings=["PATIENT NAME","PATIENT ID"],values=all_patients,hide_vertical_scroll=False,max_col_width=200,enable_events=True,key = "patient_table")],
                   [sg.B("ADD NEW PATIENT",key = "Add_patient")],
                   [sg.B("Logout",key="Logout")]]

    homewindow = sg.Window("Home",homepage_layout)
    event, values = homewindow.Read()
    while event != "Logout":
        if event in (None, 'Exit'):
            break
        if event == "Add_patient":
            homewindow.close()
            new_patient()
        else:
            print("BP1 in if event not login event",event)
            homewindow.close()
            get_patient_details(int(values["patient_table"][0]))
        event, values = homewindow.Read()
        print ("BP2 event",event,values)
    print ("BP3 event",event,values)
    print ("BP 4event",event,values)
    homewindow.close()
    
    print ("BP5 event after homewindowclose",event,values)
    
def verify(username,password):
    data = pd.read_excel("UserLogins.xlsx")
    i = 0
    xy = data.shape
    Found = False
    while i<xy[0]:
        if ((username==data.at[i,'Username'])and(password==data.at[i,'Password'])):
            Found = True
            break
        i= i+1
    return Found
def verify_id(PID):
    data = pd.read_excel("AllPatients.xlsx")
    #flag = 0
    i = 0
    xy = data.shape
    #errFound = [False,0,0,0]
    while i<xy[0]:
        if ((PID==data.at[i,'ID'])):
            return 1
        i = i+1
    return 0
def verify_register(newuser,password,confirm):
    data = pd.read_excel("UserLogins.xlsx")
    i = 0
    xy = data.shape
    errFound = [False,0,0,0]
    while i<xy[0]:
        if ((newuser==data.at[i,'Username'])):
            errFound[0] = True
            errFound[1] = 1 #username unavailable
            break
        i= i+1
    if len(password)<8:
        errFound[0] = True
        errFound[2] = 1 #password length error
    if password!=confirm:
        errFound[0] = True
        errFound[3] = 1 #password match error
    return errFound
        
login_layout = [[sg.Text('Username : '),sg.InputText(key = "username")],
            [sg.Text('Password : '), sg.InputText(key ="password",password_char="*")],
            [sg.T("Invalid Username or Password! Please Try again!",key="Validity check",visible=False)],
            [sg.Button('LOGIN',key="Login")], 
            [sg.T("New?"),sg.Button('Register here',key="Register")]]

register_layout = [[sg.Text('Enter your username : '),sg.InputText(key = "new_username")],
            [sg.T("The Username has been taken! Please Try another one!",key="usernamecheck",visible=False)],
            [sg.Text('Enter your password : '), sg.InputText(key ="new_password",password_char="*")],
            [sg.T("Password length should be atleast 8",key="Validity check",visible=False)],
            [sg.Text('Confirm your password : '), sg.InputText(key ="confirm_password",password_char="*")],
            [sg.T("Original and confirmed passwords do not match!",key="passwordcheck",visible=False)],
            [sg.Button('Submit',key="Submit")]]
data = pd.read_excel("UserLogins.xlsx")

def register():
    registerwindow = sg.Window("Register",register_layout)
    while True:    # Event Loop
        event, values = registerwindow.Read()
        if event in (None, 'Exit'):
            break
        if event == 'Submit':
            errfound = verify_register(values["new_username"],values["new_password"],values["confirm_password"])
            if (errfound[0]==False):
                print("BP13 in register",event,values)
                ip = [values["new_username"],values["new_password"]]
                wb = openpyxl.load_workbook("UserLogins.xlsx")
                r = wb.active
                r.append(ip)
                print("BP14 in register",event,values)
                wb.save("UserLogins.xlsx")
                print("BP15 in register",event,values)
                registerwindow.close()
                #login()
                print("BP16 in register",event,values)
            elif (errfound[0]==True):
                if errfound[1]==1:
                    registerwindow["usernamecheck"].update(visible = True)
                else:
                    registerwindow["usernamecheck"].update(visible = False)
                if errfound[2]==1:
                    registerwindow["Validity check"].update(visible = True)
                else:
                    registerwindow["Validity check"].update(visible = False)
                if errfound[3]==1:
                    registerwindow["passwordcheck"].update(visible = True)
                else:
                    registerwindow["passwordcheck"].update(visible = False)
            
    #registerwindow.close()
            
def login(): 
    loginwindow = sg.Window("Login",login_layout)
    while True:
        print ("BP11 event1")           # Event Loopp
        event, values = loginwindow.Read()
        print ("BP10 event1",event,values)
        if event in (None, 'Exit'):
            break
        if event == 'Login':
            #print('Pressed login')
            loginID = values['username']
            password = values['password']
            #data=[]
            #print("loginid = ",loginID,"  password = ",password)
            print("BP6 event value here in ==login loop",event,values)
            #print("event ",event)
            #print("value",values)
            check = verify(loginID,password)
            if check == False:
                #print("check",check)
                loginwindow["Validity check"].update(visible = True)
            else:
                loginwindow["Validity check"].update(visible = False)
                #print("B4 login close")
                #loginwindow.close()
                print("BP7 bfr go home event vals",event,values)
                loginwindow.Hide()
                go_to_home()
                #sg.popup("Logged Out!")
                loginwindow['password'].update(value="")
                loginwindow.UnHide()
                print("BP8 After goto event",event)
                #print("After calling go to home")
        elif event == 'Register':
            print('Pressed register')
            loginwindow.Hide()
            register()
            loginwindow['username'].update(value="")
            loginwindow['password'].update(value="")
            loginwindow.UnHide()
#FinalANN.funcpredict()
login()
print("THE END")